from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Dict, Optional, Tuple
from io import BytesIO
from openpyxl import load_workbook

app = FastAPI(title="Excel Template Merge Service")

class TableConfig(BaseModel):
    sheet_target: str
    header_row: int = 1
    start_row_target: int
    start_col_target: int
    columns: Dict[str, str]          # header_sorgente -> header_logico_template (usato solo per ordine)
    max_rows: Optional[int] = None

class MergeConfig(BaseModel):
    single_fields_by_header: Dict[str, Tuple[str, str]]        # "HeaderBrutto": ["FoglioTgt","B5"]
    table_mappings: Dict[str, TableConfig]                     # "FoglioBrutto": TableConfig
    single_fields_by_cell: Optional[Dict[str, Tuple[str, str]]] = None  # "FoglioSrc!B3": ["FoglioTgt","E9"]

def _header_index_map(ws, header_row=1):
    idx = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None and str(v).strip():
            idx[str(v).strip()] = c
    return idx

@app.get("/")
def health():
    return {"status": "ok"}

@app.post("/merge")
async def merge_excel(
    template_file: UploadFile = File(...),  # template originale (xlsx/xlsm) pre-conversione
    ugly_file: UploadFile = File(...),      # file “brutto” post-logica (xlsx)
    config_json: str = Form(...),           # stringa JSON conforme a MergeConfig
):
    # Controlli basilari formato
    tpl_name = (template_file.filename or "").lower()
    if not (tpl_name.endswith(".xlsx") or tpl_name.endswith(".xlsm")):
        raise HTTPException(400, "Il template deve essere .xlsx o .xlsm")

    try:
        cfg = MergeConfig.model_validate_json(config_json)
    except Exception as e:
        raise HTTPException(400, f"config_json non valido: {e}")

    tpl_bytes = BytesIO(await template_file.read())
    ugly_bytes = BytesIO(await ugly_file.read())

    # Apri template preservando macro se presenti
    try:
        wb_tgt = load_workbook(tpl_bytes, data_only=False, keep_vba=True)
    except Exception as e:
        raise HTTPException(400, f"Template non apribile: {e}")

    # Apri sorgente dati (ugly.xlsx) – solo valori calcolati
    try:
        wb_src = load_workbook(ugly_bytes, data_only=True, keep_vba=True)
    except Exception as e:
        raise HTTPException(400, f"Ugly non apribile: {e}")

    # 1) Campi singoli per header (assume valore sotto l'header)
    header_values = {}
    for ws in wb_src.worksheets:
        hdr = _header_index_map(ws, header_row=1)
        for h, c in hdr.items():
            val = ws.cell(row=2, column=c).value  # adatta se la tua logica è diversa
            header_values[h] = val

    for header, (tgt_sheet, tgt_cell) in cfg.single_fields_by_header.items():
        if header not in header_values:
            raise HTTPException(400, f"Header '{header}' non trovato nell'ugly.")
        wb_tgt[tgt_sheet][tgt_cell].value = header_values[header]

    # 2) Campi singoli cella→cella
    if cfg.single_fields_by_cell:
        for src_key, (tgt_sheet, tgt_cell) in cfg.single_fields_by_cell.items():
            if "!" not in src_key:
                raise HTTPException(400, f"Chiave single_fields_by_cell non valida: {src_key}")
            src_sheet, src_cell = src_key.split("!", 1)
            try:
                value = wb_src[src_sheet][src_cell].value
            except KeyError:
                raise HTTPException(400, f"Sorgente non trovata: {src_key}")
            wb_tgt[tgt_sheet][tgt_cell].value = value

    # 3) Tabelle per header
    for src_sheet, tcfg in cfg.table_mappings.items():
        if src_sheet not in wb_src.sheetnames:
            raise HTTPException(400, f"Foglio sorgente '{src_sheet}' non trovato.")
        ws_src = wb_src[src_sheet]
        if tcfg.sheet_target not in wb_tgt.sheetnames:
            raise HTTPException(400, f"Foglio target '{tcfg.sheet_target}' non trovato.")
        ws_tgt = wb_tgt[tcfg.sheet_target]

        src_hdr_idx = _header_index_map(ws_src, header_row=tcfg.header_row)
        ordered_src_cols = []
        for h_src in tcfg.columns.keys():
            if h_src not in src_hdr_idx:
                raise HTTPException(400, f"Colonna '{h_src}' non trovata nel foglio '{src_sheet}'.")
            ordered_src_cols.append(src_hdr_idx[h_src])

        data_start = tcfg.header_row + 1
        written = 0
        r_src = data_start
        while True:
            row_vals = [ws_src.cell(row=r_src, column=c).value for c in ordered_src_cols]
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in row_vals):
                break
            if tcfg.max_rows is not None and written >= tcfg.max_rows:
                break

            c_tgt = tcfg.start_col_target
            for v in row_vals:
                ws_tgt.cell(row=tcfg.start_row_target + written, column=c_tgt).value = v
                c_tgt += 1

            written += 1
            r_src += 1

    # Salva in memoria e restituisci
    out_bytes = BytesIO()
    wb_tgt.save(out_bytes)  # preserva macro se c’erano (.xlsm)
    out_bytes.seek(0)

    filename = "output_finale.xlsm" if tpl_name.endswith(".xlsm") else "output_finale.xlsx"
    media = (
        "application/vnd.ms-excel.sheet.macroEnabled.12"
        if filename.endswith(".xlsm")
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    return StreamingResponse(
        out_bytes,
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
